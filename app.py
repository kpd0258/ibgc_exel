import os
import requests
from datetime import datetime, timezone, timedelta, date
from flask import Flask, jsonify, request, send_from_directory
from flask_cors import CORS
from openpyxl import load_workbook

app = Flask(__name__)
CORS(app)

# =========================
# ENV VARIABLES
# =========================

EXCEL_API_KEY = os.environ.get("EXCEL_API_KEY", "")

BUBBLE_BASE_URL = os.environ.get("BUBBLE_BASE_URL", "").rstrip("/")  # ex) https://ibgc.co.kr/version-test
BUBBLE_DATA_API_TOKEN = os.environ.get("BUBBLE_DATA_API_TOKEN", "")
BUBBLE_APP_TYPE = os.environ.get("BUBBLE_APP_TYPE", "00. Application")
TEMPLATE_PATH = os.environ.get("TEMPLATE_PATH", "IBGC_Application_Template.xlsx")

# Bubble Data API base
BUBBLE_DATA_API_BASE = f"{BUBBLE_BASE_URL}/api/1.1/obj"

# Bubble fileupload endpoint
BUBBLE_FILEUPLOAD_URL = os.environ.get("BUBBLE_FILEUPLOAD_URL", f"{BUBBLE_BASE_URL}/fileupload")

KST = timezone(timedelta(hours=9))

# =========================
# SHEET NAMES (4 sheets)
# =========================
SHEET_1 = "1. 인증신청서 관리"
SHEET_2 = "2. 인결위 일정 관리"
SHEET_3 = "3. UAF_IAF"
SHEET_4 = "4. 인증서 발행 관리대장=홈페이지 db+우편"

# Related data type names in Bubble (for reference fields)
TYPE_APP_FORM = "01. Application Form"
TYPE_CALCU_DB = "03. Calcu_DB"
TYPE_AUDIT_PROGRAM = "06. Audit Program"
TYPE_DECISION_STAGE1 = "08. Decision Report : Stage1"
TYPE_DECISION_STAGE2 = "09. Decision Report : Stage2"
TYPE_CERT_ISSUANCE = "11. Certificate issuance"
TYPE_CHAPTER_VIEWER = "Chapter Viewer"  # ✅ 추가


# =========================
# UTIL
# =========================

def require_api_key(req):
    if not EXCEL_API_KEY:
        return True
    return req.headers.get("X-API-Key") == EXCEL_API_KEY


def now_kst():
    return datetime.now(KST)


def today_label():
    # include time to avoid duplicates
    return now_kst().strftime("IBGC_Application_%Y%m%d_%H%M%S.xlsx")


def bubble_headers():
    return {
        "Authorization": f"Bearer {BUBBLE_DATA_API_TOKEN}",
        "Content-Type": "application/json",
    }


def ok_status(code: int) -> bool:
    # Bubble create=201, update=204 are common
    return code in (200, 201, 204)


def safe_str(v, default=""):
    if v is None:
        return default
    s = str(v).strip()
    return s if s != "" else default


def parse_bubble_date(value):
    """
    Bubble can return:
      - ISO string: "2026-03-02T12:34:56.789Z"
      - ISO string without Z
      - date-only string: "2026-03-02"
      - already datetime/date
      - None
    Returns datetime|date|None
    """
    if value is None or value == "":
        return None
    if isinstance(value, datetime) or isinstance(value, date):
        return value
    if not isinstance(value, str):
        return None

    s = value.strip()
    if s == "":
        return None

    # normalize Z
    try:
        if s.endswith("Z"):
            # fromisoformat doesn't accept Z, convert to +00:00
            s2 = s[:-1] + "+00:00"
            return datetime.fromisoformat(s2)
        # date-only
        if len(s) == 10 and s[4] == "-" and s[7] == "-":
            return datetime.fromisoformat(s).date()
        return datetime.fromisoformat(s)
    except Exception:
        return None


def fmt_yyyy_mm_dd(value, fallback="미해당"):
    d = parse_bubble_date(value)
    if d is None:
        return fallback
    if isinstance(d, datetime):
        return d.astimezone(KST).date().strftime("%Y-%m-%d")
    if isinstance(d, date):
        return d.strftime("%Y-%m-%d")
    return fallback


def fmt_month_eng_d_yyyy(value, fallback="미해당"):
    d = parse_bubble_date(value)
    if d is None:
        return fallback
    if isinstance(d, datetime):
        d = d.astimezone(KST).date()
    if isinstance(d, date):
        # e.g., "March 2 2026"
        return d.strftime("%B %-d %Y") if os.name != "nt" else d.strftime("%B %#d %Y")
    return fallback


def fmt_1_decimal(value, fallback=""):
    if value is None or value == "":
        return fallback
    try:
        return f"{float(value):.1f}"
    except Exception:
        return fallback


def normalize_standards(value):
    """
    cert_standards might be:
      - list of strings
      - comma-separated string
      - bubble option set text
    Return set of normalized tokens.
    """
    if value is None:
        return set()
    if isinstance(value, list):
        items = [safe_str(x) for x in value]
    else:
        s = safe_str(value)
        if s == "":
            return set()
        # split by comma
        items = [x.strip() for x in s.split(",")]
    # normalize
    out = set()
    for it in items:
        if it:
            out.add(it.upper())
    return out


def iso_combo_label(standards_set):
    q = "ISO 9001" in standards_set
    e = "ISO 14001" in standards_set
    o = "ISO 45001" in standards_set

    if q and e and o:
        return "QEO"
    if q and e and not o:
        return "QE(9001+14001)"
    if e and o and not q:
        return "EO(14001+45001)"
    if q and o and not e:
        return "QO(9001+45001)"
    if q and not e and not o:
        return "Q(9001)"
    if e and not q and not o:
        return "E(14001)"
    if o and not q and not e:
        return "O(45001)"
    return ""


def map_application_type3(value):
    v = safe_str(value, "")
    if v == "최초":
        return "Initial"
    if v == "전환":
        return "Transferred"
    if v == "특별":
        return "Special"
    return v  # if already English or other


def first_nonempty(*vals, default="미해당"):
    for v in vals:
        s = safe_str(v, "")
        if s != "":
            return s
    return default


# =========================
# BUBBLE DATA API HELPERS
# =========================

_related_cache = {}

def bubble_get_object(type_name, obj_id):
    if not obj_id:
        return None
    key = (type_name, obj_id)
    if key in _related_cache:
        return _related_cache[key]

    url = f"{BUBBLE_DATA_API_BASE}/{type_name}/{obj_id}"
    res = requests.get(url, headers=bubble_headers(), timeout=60)
    if not ok_status(res.status_code):
        # cache as None to avoid hammering
        _related_cache[key] = None
        return None

    data = res.json().get("response", {})
    _related_cache[key] = data
    return data


def resolve_related(app_data, field_key, type_name):
    """
    app_data[field_key] could be:
      - dict (already expanded)
      - string (thing id)
      - None
    """
    v = app_data.get(field_key)
    if isinstance(v, dict):
        return v
    if isinstance(v, str) and v.strip():
        return bubble_get_object(type_name, v.strip())
    return None


def get_all_applications():
    url = f"{BUBBLE_DATA_API_BASE}/{BUBBLE_APP_TYPE}"
    res = requests.get(url, headers=bubble_headers(), timeout=90)
    if not ok_status(res.status_code):
        raise Exception(f"Bubble fetch error ({res.status_code}): {res.text}")
    return res.json().get("response", {}).get("results", [])


def create_daily_excel_record(file_bubble_url, file_url, label, status="ready", source_count=0, note=""):
    url = f"{BUBBLE_DATA_API_BASE}/DailyExcel"
    payload = {
        "file": file_bubble_url,
        "file_url": file_url,
        "label": label,
        "status": status,
        "source_count": source_count,
        "note": note
    }
    res = requests.post(url, headers=bubble_headers(), json=payload, timeout=60)
    if not ok_status(res.status_code):
        raise Exception(f"Bubble create error ({res.status_code}): {res.text}")
    return res.json() if res.text else {"status": "success"}


def upload_file_to_bubble_storage(local_path: str, upload_filename: str) -> str:
    if not os.path.exists(local_path):
        raise Exception(f"File not found: {local_path}")

    with open(local_path, "rb") as f:
        files = {
            "file": (upload_filename, f, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        }
        res = requests.post(BUBBLE_FILEUPLOAD_URL, files=files, timeout=180)

    if not ok_status(res.status_code):
        raise Exception(f"Bubble fileupload error ({res.status_code}): {res.text}")

    text = (res.text or "").strip()

    # JSON case (rare)
    if text.startswith("{"):
        try:
            j = res.json()
            if "url" in j:
                return j["url"]
        except Exception:
            pass

    # Common case: raw URL string
    if text.startswith("//") or text.startswith("http"):
        return text

    # quoted URL
    if text.startswith('"//') and text.endswith('"'):
        return text.strip('"')

    raise Exception(f"Unexpected fileupload response: {res.text}")


# =========================
# EXCEL GENERATION (Sheet 1 filled as requested)
# =========================

def fill_sheet_1(ws, applications):
    """
    Sheet 1: "1. 인증신청서 관리"
    start row: 9
    Col mapping: 1..78
    """
    start_row = 9
    row = start_row

    for idx, app_data in enumerate(applications, start=1):

        # resolve referenced things (if Bubble returns ids)
        app_form = resolve_related(app_data, "01. Application Form", TYPE_APP_FORM) or {}
        calcu_db = resolve_related(app_data, "03. Calcu_DB", TYPE_CALCU_DB) or {}
        audit_program = resolve_related(app_data, "06. Audit Program", TYPE_AUDIT_PROGRAM) or {}
        dr_stage1 = resolve_related(app_data, "08. Decision Report : Stage1", TYPE_DECISION_STAGE1) or {}
        dr_stage2 = resolve_related(app_data, "09. Decision Report : Stage2", TYPE_DECISION_STAGE2) or {}
        cert_issuance = resolve_related(app_data, "11. Certificate issuance", TYPE_CERT_ISSUANCE) or {}

        # ✅ NEW: Chapter Viewer
        chapter_viewer = resolve_related(app_data, "Chapter Viewer", TYPE_CHAPTER_VIEWER) or {}

        # standards flags
        standards_set = normalize_standards(app_data.get("cert_standards"))
        has_9001 = "ISO 9001" in standards_set
        has_14001 = "ISO 14001" in standards_set
        has_45001 = "ISO 45001" in standards_set

        # 1..78 columns
        c = {}

        # 1 index
        c[1] = idx

        # 2 in_charge_of
        c[2] = safe_str(app_data.get("in_charge_of"), "")

        # 3 recommend (default IBGC)
        c[3] = safe_str(app_data.get("recommend"), "IBGC")

        # 4 JOB_NO
        c[4] = safe_str(app_data.get("JOB_NO"), "")

        # 5 "O"
        c[5] = "O"

        # 6 Org Name (kor)
        c[6] = safe_str(app_form.get("Organization Name(kor)"), "")

        # 7 Org Name (eng)
        c[7] = safe_str(app_form.get("Organization Name(eng)"), "")

        # 8 ISO 9001 => 1 else "-"
        c[8] = "1" if has_9001 else "-"

        # 9 ISO 14001 => 1 else "-"
        c[9] = "1" if has_14001 else "-"

        # 10 ISO 45001 => 1 else "-"
        c[10] = "1" if has_45001 else "-"

        # 11 application_type(3) map
        c[11] = map_application_type3(app_data.get("application_type(3)"))

        # 12 application_type(5)
        c[12] = safe_str(app_data.get("application_type(5)"), "")

        # 13 combo label
        c[13] = iso_combo_label(standards_set)

        # 14 Certification Scope(eng)
        c[14] = safe_str(app_form.get("Certification Scope(eng)"), "")

        # 15 Certification Scope(kor)
        c[15] = safe_str(app_form.get("Certification Scope(kor)"), "")

        # 16 IAF CODE
        c[16] = safe_str(app_form.get("IAF CODE"), "")

        # 17 blank
        c[17] = ""

        # 18 EMS RISK
        c[18] = safe_str(app_form.get("EMS RISK"), "")

        # 19 OHS RISK
        c[19] = safe_str(app_form.get("OHS RISK"), "")

        # 20 Business Registration No.
        c[20] = safe_str(app_form.get("Business Registration No."), "")

        # 21 President(kor)
        c[21] = safe_str(app_form.get("President(kor)"), "")

        # 22 Contact Person (on Application)
        c[22] = safe_str(app_data.get("Contact Person"), "")

        # 23 Contact person's Tel
        c[23] = safe_str(app_form.get("Contact person's Tel"), "")

        # 24 Contact person's E-mail
        c[24] = safe_str(app_form.get("Contact person's E-mail"), "")

        # 25 Address kor full + ", " + detail(kor)
        addr_full_kor = safe_str(app_form.get("Organization_adress_full(kor)"), "")
        addr_detail_kor = safe_str(app_form.get("Organization_adress_detail(kor)"), "")
        c[25] = (addr_full_kor + (", " if addr_full_kor and addr_detail_kor else "") + addr_detail_kor).strip()

        # 26 blank
        c[26] = ""

        # 27 detail(eng) + ", " + full(eng)
        addr_detail_eng = safe_str(app_form.get("Organization_adress_detail(eng)"), "")
        addr_full_eng = safe_str(app_form.get("Organization_adress_full(eng)"), "")
        c[27] = (addr_detail_eng + (", " if addr_detail_eng and addr_full_eng else "") + addr_full_eng).strip()

        # 28 "(" + postcode + ")" + full(eng) + ", " + detail(eng)
        postcode = safe_str(app_form.get("Organization_postcode"), "")
        core = addr_full_eng + (", " if addr_full_eng and addr_detail_eng else "") + addr_detail_eng
        if postcode and core:
            c[28] = f"({postcode}){core}"
        elif postcode and not core:
            c[28] = f"({postcode})"
        else:
            c[28] = core.strip()

        # 29 postcode
        c[29] = postcode

        # 30 Outsourcing process default none
        c[30] = safe_str(app_form.get("Outsourcing process"), "none")

        # 31 Number of Employees(certi)
        c[31] = safe_str(app_form.get("Number of Employees(certi)"), "")

        # 32 Name of Product/Service1
        c[32] = safe_str(app_form.get("Name of Product/Service1"), "")

        # 33 Recently Date of audit (MonthEng day year) else 미해당
        c[33] = fmt_month_eng_d_yyyy(app_form.get("Recently Date of audit"), fallback="미해당")

        # 34 Recently audit type(5) fallback chain else 미해당
        c[34] = first_nonempty(
            app_form.get("Recently_9001_Audit Type(5)"),
            app_form.get("Recently_14001_Audit Type(5)"),
            app_form.get("Recently_45001_Audit Type(5)"),
            default="미해당"
        )

        # 35 Next Date of audit
        c[35] = fmt_month_eng_d_yyyy(app_form.get("Next Date of audit"), fallback="미해당")

        # 36 Next audit type(5) fallback chain
        c[36] = first_nonempty(
            app_form.get("Next_9001_Audit Type(5)"),
            app_form.get("Next_14001_Audit Type(5)"),
            app_form.get("Next_45001_Audit Type(5)"),
            default="미해당"
        )

        # 37 _인증전환신청_인증서유효기간 yyyy-mm-dd else 미해당
        c[37] = fmt_yyyy_mm_dd(app_data.get("_인증전환신청_인증서유효기간"), fallback="미해당")

        # 38 _인증전환신청_타인증기관명 else 미해당
        c[38] = safe_str(app_data.get("_인증전환신청_타인증기관명"), "미해당")

        # 39 _법적의무사항_관련법
        c[39] = safe_str(app_data.get("_법적의무사항_관련법"), "")

        # 40 Declaration of date yyyy-mm-dd
        c[40] = fmt_yyyy_mm_dd(app_form.get("Declaration of date"), fallback="")

        # 41 stage1 audit date yyyy-mm-dd
        c[41] = fmt_yyyy_mm_dd(app_form.get("stage1 audit date"), fallback="")

        # 42 stage2 audit date yyyy-mm-dd
        c[42] = fmt_yyyy_mm_dd(app_form.get("stage2 audit date"), fallback="")

        # 43 Final M/D-stage1 (1 decimal)
        c[43] = fmt_1_decimal(calcu_db.get("17. Final M/D-stage1"), fallback="")

        # 44 Final M/D-stage2 (1 decimal)
        c[44] = fmt_1_decimal(calcu_db.get("17. Final M/D-stage2"), fallback="")

        # 45 Final M/D-eachsurv (1 decimal)
        c[45] = fmt_1_decimal(calcu_db.get("17. Final M/D-eachsurv"), fallback="")

        # 46 Final M/D-recert (1 decimal)
        c[46] = fmt_1_decimal(calcu_db.get("17. Final M/D-recert"), fallback="")

        # 47 stage1 + stage2 (1 decimal)
        try:
            v1 = float(calcu_db.get("17. Final M/D-stage1") or 0)
            v2 = float(calcu_db.get("17. Final M/D-stage2") or 0)
            c[47] = f"{(v1 + v2):.1f}"
        except Exception:
            c[47] = ""

        # ✅ 48: '김정석' (고정)
        c[48] = "김정석"

        # ✅ 49: Chapter Viewer - Technical reviewer
        c[49] = safe_str(chapter_viewer.get("Technical reviewer"), "")

        # ✅ 50: Chapter Viewer - Date of Receipt (yyyy-mm-dd)
        c[50] = fmt_yyyy_mm_dd(chapter_viewer.get("Date of Receipt"), fallback="")

        # 51 Lead auditor
        c[51] = safe_str(app_form.get("Lead auditor"), "")

        # 52 auditor
        c[52] = safe_str(app_form.get("auditor"), "")

        # 53 provisional auditor
        c[53] = safe_str(app_form.get("provisional auditor"), "")

        # 54 technical expert
        c[54] = safe_str(app_form.get("technical expert"), "")

        # 55 observer
        c[55] = safe_str(app_form.get("observer"), "")

        # ✅ 56: '김정석' (고정)
        c[56] = "김정석"

        # 57 Date of Audit_transfer yyyy-mm-dd
        c[57] = fmt_yyyy_mm_dd(audit_program.get("Date of Audit_transfer"), fallback="")

        # 58 Date of Audit_Re_certi yyyy-mm-dd
        c[58] = fmt_yyyy_mm_dd(audit_program.get("Date of Audit_Re_certi"), fallback="")

        # 59 Date of Audit_select1 yyyy-mm-dd
        c[59] = fmt_yyyy_mm_dd(audit_program.get("Date of Audit_select1"), fallback="")

        # 60 Date of Audit_select2 yyyy-mm-dd
        c[60] = fmt_yyyy_mm_dd(audit_program.get("Date of Audit_select2"), fallback="")

        # 61 Date of Audit_select3 yyyy-mm-dd
        c[61] = fmt_yyyy_mm_dd(audit_program.get("Date of Audit_select3"), fallback="")

        # 62 Audit_md_01
        c[62] = safe_str(audit_program.get("Audit_md_01"), "")

        # 63 Audit_md_02
        c[63] = safe_str(audit_program.get("Audit_md_02"), "")

        # 64 Audit_md_03
        c[64] = safe_str(audit_program.get("Audit_md_03"), "")

        # 65 Audit_md_04
        c[65] = safe_str(audit_program.get("Audit_md_04"), "")

        # 66 Audit_md_05
        c[66] = safe_str(audit_program.get("Audit_md_05"), "")

        # ✅ 67: Stage1 audit type(5) "있으면" Stage 1 Certification audit, 없으면 '-'
        c[67] = "Stage 1 Certification audit" if safe_str(dr_stage1.get("audit type(5)"), "") != "" else "-"

        # ✅ 68: Stage1 Approval of Evaluation_Date yyyy-mm-dd else '-'
        c[68] = fmt_yyyy_mm_dd(dr_stage1.get("Approval of Evaluation_Date"), fallback="-")

        # ✅ 69: Stage2 audit type(5) "있으면" Stage 2 Certification audit, 없으면 '-'
        c[69] = "Stage 2 Certification audit" if safe_str(dr_stage2.get("audit type(5)"), "") != "" else "-"

        # ✅ 70: Stage2 Verification Date yyyy-mm-dd else '-'
        c[70] = fmt_yyyy_mm_dd(dr_stage2.get("Verification Date"), fallback="-")

        # 71 Reviewer of Certification Records else '-'
        c[71] = safe_str(dr_stage2.get("Reviewer of Certification Records"), "-")

        # 72 constant
        c[72] = "GM 이원호"

        # 73 Initial Date_* yyyy-mm-dd (first among 9001/14001/45001)
        c[73] = fmt_yyyy_mm_dd(
            first_nonempty(
                cert_issuance.get("Initial Date_9001"),
                cert_issuance.get("Initial Date_14001"),
                cert_issuance.get("Initial Date_45001"),
                default=""
            ),
            fallback=""
        )

        # 74 Issue Date_* yyyy-mm-dd
        c[74] = fmt_yyyy_mm_dd(
            first_nonempty(
                cert_issuance.get("Issue Date_9001"),
                cert_issuance.get("Issue Date_14001"),
                cert_issuance.get("Issue Date_45001"),
                default=""
            ),
            fallback=""
        )

        # 75 Expire Date_* yyyy-mm-dd
        c[75] = fmt_yyyy_mm_dd(
            first_nonempty(
                cert_issuance.get("Expire Date_9001"),
                cert_issuance.get("Expire Date_14001"),
                cert_issuance.get("Expire Date_45001"),
                default=""
            ),
            fallback=""
        )

        # 76 Certificate No._9001
        c[76] = safe_str(cert_issuance.get("Certificate No._9001"), "")

        # 77 Certificate No._14001
        c[77] = safe_str(cert_issuance.get("Certificate No._14001"), "")

        # 78 Certificate No._45001
        c[78] = safe_str(cert_issuance.get("Certificate No._45001"), "")

        # write to worksheet (A=1 ..)
        for col_idx in range(1, 79):
            ws.cell(row=row, column=col_idx, value=c.get(col_idx, ""))

        row += 1


def generate_excel_file():
    applications = get_all_applications()

    if not os.path.exists(TEMPLATE_PATH):
        raise Exception(f"Template file not found: {TEMPLATE_PATH}")

    wb = load_workbook(TEMPLATE_PATH)

    # Ensure 4 sheets exist by name
    for name in (SHEET_1, SHEET_2, SHEET_3, SHEET_4):
        if name not in wb.sheetnames:
            raise Exception(f"Template missing sheet: '{name}'")

    # ✅ Fill ONLY sheet 1 for now (as requested)
    ws1 = wb[SHEET_1]
    fill_sheet_1(ws1, applications)

    filename = today_label()
    generated_dir = "generated"
    os.makedirs(generated_dir, exist_ok=True)

    file_path = os.path.join(generated_dir, filename)
    wb.save(file_path)

    # ✅ Upload to Bubble storage
    bubble_file_url = upload_file_to_bubble_storage(file_path, filename)

    if bubble_file_url.startswith("//"):
        downloadable_url = "https:" + bubble_file_url
    else:
        downloadable_url = bubble_file_url

    return {
        "local_path": file_path,
        "filename": filename,
        "bubble_file_url": bubble_file_url,
        "download_url": downloadable_url,
        "source_count": len(applications),
    }


# =========================
# ROUTES
# =========================

@app.route("/health", methods=["GET"])
def health():
    return jsonify({"ok": True})


@app.route("/excel/generate_daily", methods=["POST"])
def excel_generate_daily():
    if not require_api_key(request):
        return jsonify({"ok": False, "error": "Unauthorized"}), 401

    try:
        result = generate_excel_file()

        create_daily_excel_record(
            file_bubble_url=result["bubble_file_url"],
            file_url=result["download_url"],
            label=result["filename"],
            status="ready",
            source_count=result["source_count"],
            note=""
        )

        return jsonify({
            "ok": True,
            "file": result["bubble_file_url"],
            "file_url": result["download_url"],
            "label": result["filename"],
            "status": "ready",
            "source_count": result["source_count"],
            "created_at": now_kst().isoformat()
        })

    except Exception as e:
        return jsonify({"ok": False, "error": str(e)}), 500


@app.route("/excel/refresh_now", methods=["POST"])
def excel_refresh_now():
    return excel_generate_daily()


@app.route("/download/<filename>", methods=["GET"])
def download_local_file(filename):
    generated_dir = "generated"
    return send_from_directory(generated_dir, filename, as_attachment=True)


# =========================
# MAIN
# =========================

if __name__ == "__main__":
    app.run(host="0.0.0.0", port=int(os.environ.get("PORT", 10000)))
